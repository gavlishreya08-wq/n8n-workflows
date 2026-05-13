from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime
import sys

file_path = sys.argv[1]
wb = load_workbook(file_path)
ws = wb.active

# === Styles ===
light_red_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # < 5.5 hrs
dark_red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")   # holidays / Sundays
red_font = Font(color="FF0000")
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

# === Predefined holidays (DD-MM-YYYY format) ===
holidays = {"20-10-2025", "02-10-2025"}

# === Columns that should NOT be colored ===
skip_headers = {
    "Sr. No.",
    "Total Logged",
    "Working Hours",
    "Leave",
    "Leave Hours",
    "Actual Hours",
    "Hours Not Logged",
    "Days Not Logged"
}

# === Add Sr. No. column if missing ===
if ws.cell(row=1, column=1).value != "Sr. No.":
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="Sr. No.").font = bold_font
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
        row[0].value = i  # assign Sr. No.

# === Bold + borders for header ===
header_map = {}
for cell in ws[1]:
    val = str(cell.value).strip() if cell.value else ""
    header_map[cell.column] = val
    cell.font = bold_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# === Identify which columns are date columns ===
date_columns = {}
for cell in ws[1]:
    val = str(cell.value).strip() if cell.value else ""
    for fmt in ("%d-%b-%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            date_obj = datetime.strptime(val, fmt)
            date_columns[cell.column] = date_obj
            break
        except Exception:
            continue

# === Apply formatting ===
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border = thin_border
        header_text = header_map.get(cell.column, "")

        # Skip summary and Sr. No. columns
        if header_text in skip_headers:
            continue

        # Light red + red text if worklog < 5.5
        if isinstance(cell.value, (int, float)) and cell.value < 5.5:
            cell.fill = light_red_fill
            cell.font = red_font

        # Dark red for holidays or Sundays (no white bold text)
        col_date = date_columns.get(cell.column)
        if col_date:
            date_str = col_date.strftime("%d-%m-%Y")
            if date_str in holidays or col_date.weekday() == 6:
                cell.fill = dark_red_fill
                cell.font = Font(color="000000")  # keep text black

# === Adjust column widths ===
for col in ws.columns:
    col_letter = col[0].column_letter
    header = str(ws.cell(row=1, column=col[0].column).value or "")
    if header == "Sr. No.":
        ws.column_dimensions[col_letter].width = 7
    elif header in skip_headers:
        ws.column_dimensions[col_letter].width = 12
    elif header in ["EmployeeName", "Designation"]:
        ws.column_dimensions[col_letter].width = 22
    else:
        ws.column_dimensions[col_letter].width = 6  # compact for dates

# === Save file ===
wb.save(file_path)
