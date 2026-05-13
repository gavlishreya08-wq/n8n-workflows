from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys

file_path = sys.argv[1]
wb = load_workbook(file_path)
ws = wb.active

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Color Sundays (column names are dates)
for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
    for cell in col:
        if isinstance(cell.value, (int, float)) and cell.value < 5.5:
            cell.fill = red_fill
        if isinstance(cell.value, str) and "Sun" in cell.value:
            cell.fill = red_fill

wb.save(file_path)
