#!/usr/bin/env python3
"""
Full Python script — Option A (title row + grouping row).
- Reads input headers from ROW 1 (this matches your n8n output).
- Detects months from header names (e.g., Planned_Apr25-Oct25, Planned_Nov25).
- Group1 = all months except last chronological month (divisor 24).
- Group2 = last chronological month (divisor 19).
- Calculates Extra Persons and formats sheet, inserting a title row and a merged section header row.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import sys
import re
from datetime import datetime

if len(sys.argv) < 2:
    print("Usage: python excel_formatter_dynamic_full.py <path_to_excel_file>")
    sys.exit(1)

file_path = sys.argv[1]
wb = load_workbook(file_path)
ws = wb.active

# === Styles ===
bold_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color="000000")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

planned_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
actual_fill = PatternFill(start_color="CFE2F3", fill_type="solid")
dev_fill = PatternFill(start_color="F9CB9C", fill_type="solid")
percent_fill = PatternFill(start_color="D9EAD3", fill_type="solid")

title_fill = PatternFill(start_color="FFF59D", fill_type="solid")
extra_fill = PatternFill(start_color="FF0000", fill_type="solid")


def border_merged(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border


# -------------------------------------------------------
# Helpers: month parsing + labeling
# -------------------------------------------------------
MONTHS = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}

month_pattern_two = re.compile(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[\s\-]?(\d{2})', re.IGNORECASE)
month_pattern_any = re.compile(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)', re.IGNORECASE)

def make_label_from_tuple(t):
    # t is (year_full, monthnum, label)
    return t[2]

def make_group_label(months):
    # months: list of (year, monthnum, label)
    if not months:
        return ""
    if len(months) == 1:
        return months[0][2]
    start_label = months[0][2]
    end_label = months[-1][2]
    return f"{start_label}-{end_label}"


# -------------------------------------------------------
# STEP A: Build header_to_col from INPUT header row (row 1)
# This is the row produced by your JS (n8n) node.
# -------------------------------------------------------
HEADER_INPUT_ROW = 1

header_to_col = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=HEADER_INPUT_ROW, column=c).value
    if v is not None:
        header_to_col[str(v)] = c

# -------------------------------------------------------
# STEP B: Detect months from headers (robust)
# - First try to find two-digit year tokens (Apr25 etc.)
# - If none, try to extract month names (use current year)
# -------------------------------------------------------
found_months = []

def extract_months_from_text(text):
    out = []
    if not text:
        return out
    text_s = str(text)
    # find all two-digit tokens
    for m in month_pattern_two.findall(text_s):
        mon, yy = m
        mon3 = mon[:3].title()
        mon_num = MONTHS.get(mon3.lower())
        year_full = 2000 + int(yy)
        label = f"{mon3}{yy}"
        out.append((year_full, mon_num, label))
    # if none, fallback to month names (use current year)
    if not out:
        for m in month_pattern_any.findall(text_s):
            mon = m[:3].title()
            mon_num = MONTHS.get(mon.lower())
            year_full = datetime.now().year
            yy = str(year_full)[-2:]
            label = f"{mon}{yy}"
            out.append((year_full, mon_num, label))
    return out

# gather from all headers on input row
for header_text in header_to_col.keys():
    extracted = extract_months_from_text(header_text)
    for t in extracted:
        found_months.append(t)

# deduplicate by (year, month) key and sort
unique = {}
for y, mnum, label in found_months:
    unique[(y, mnum)] = (y, mnum, label)

months_detected = sorted(unique.values(), key=lambda x: (x[0], x[1]))

# Build group1 (all except last), group2 (last)
if months_detected:
    if len(months_detected) >= 2:
        group1_months = months_detected[:-1]
        group2_months = [months_detected[-1]]
    else:
        group1_months = []
        group2_months = [months_detected[-1]]
else:
    group1_months = []
    group2_months = []

group1_label = make_group_label(group1_months)
group2_label = make_group_label(group2_months)

# If group1 empty keep placeholder so column names do not collide with blank
if not group1_label and group2_label:
    group1_label = f"{group2_label}-prev"

# Build canonical header keys we will use in script
group_keys = {
    'Planned_G1': f"Planned_{group1_label}",
    'Planned_G2': f"Planned_{group2_label}",
    'Planned_Total': "Planned_Total",

    'Actual_G1': f"Actual_{group1_label}",
    'Actual_G2': f"Actual_{group2_label}",
    'Actual_Total': "Actual_Total",

    'Dev_G1': f"Dev_{group1_label}",
    'Dev_G2': f"Dev_{group2_label}",
    'Dev_Total': "Dev_Total",

    'Dev%_G1': f"Dev%_{group1_label}",
    'Dev%_G2': f"Dev%_{group2_label}",
    'Dev%_Total': "Dev%_Total"
}

# -------------------------------------------------------
# STEP C: Identify Dev_Total column BEFORE sorting (from input headers)
# -------------------------------------------------------
dev_total_col = None
for c in range(1, ws.max_column + 1):
    if ws.cell(row=HEADER_INPUT_ROW, column=c).value == "Dev_Total":
        dev_total_col = c
        break

if dev_total_col is None:
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=HEADER_INPUT_ROW, column=c).value or "") == "Dev_Total":
            dev_total_col = c
            break

# fallback if not found
if dev_total_col is None:
    # try to find Dev_Total by dynamic name (if present)
    try:
        dev_total_col = header_to_col.get("Dev_Total") or header_to_col.get(group_keys['Dev_Total']) or 1
    except:
        dev_total_col = 1

# -------------------------------------------------------
# STEP D: Gather data rows (from input rows starting row 2)
# and sort by Dev_Total (descending) — KEEP ORIGINAL BEHAVIOR
# -------------------------------------------------------
max_col = ws.max_column
data_rows = []

for r in range(2, ws.max_row + 1):
    row_values = []
    for c in range(1, max_col + 1):
        row_values.append(ws.cell(row=r, column=c).value)
    val = row_values[dev_total_col - 1]
    try:
        sort_val = float(val)
    except:
        sort_val = None
    data_rows.append((sort_val, row_values))

data_rows.sort(key=lambda x: (x[0] is None, -x[0] if (x[0] is not None) else 0))

write_row = 2
for _, row_values in data_rows:
    for c in range(1, max_col + 1):
        ws.cell(row=write_row, column=c).value = row_values[c - 1]
    write_row += 1

# -------------------------------------------------------
# STEP E: Insert title row and section header row (Option A)
# After this, the data headers will land at row 3
# -------------------------------------------------------
ws.insert_rows(1)
# FIXED → merge across ALL columns (ws.max_column)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)

title_cell = ws.cell(row=1, column=1, value="Deviation % more than 15%")
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = center
title_cell.fill = title_fill

# FIXED → corrected border range (ws.max_column, not ws.max_col)
border_merged(ws, f"A1:{get_column_letter(ws.max_column)}1")

ws.insert_rows(2)

# sections mapping uses the group_keys
sections = {
    "Planned Days": [group_keys['Planned_G1'], group_keys['Planned_G2'], group_keys['Planned_Total']],
    "Actual Days": [group_keys['Actual_G1'], group_keys['Actual_G2'], group_keys['Actual_Total']],
    "Deviation Days": [group_keys['Dev_G1'], group_keys['Dev_G2'], group_keys['Dev_Total']],
    "Total Deviation %": [group_keys['Dev%_G1'], group_keys['Dev%_G2'], group_keys['Dev%_Total']],
}

# Build a header_to_col_current from the NEW header row (which is now row 3)
HEADER_OUTPUT_ROW = 3
header_to_col_output = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=HEADER_OUTPUT_ROW, column=c).value
    if v:
        header_to_col_output[str(v)] = c

# fallback: if output headers are empty (rare), use previous input mapping but offsets apply
if not header_to_col_output:
    # attempt to copy input mapping but keep column numbers same
    header_to_col_output = header_to_col.copy()

# Helper to find column by target name (exact or contains token)
def find_col_for_header_name_in_output(target):
    if target in header_to_col_output:
        return header_to_col_output[target]
    token = target.split('_', 1)[1] if '_' in target else target
    for k, col in header_to_col_output.items():
        if token and token in str(k):
            return col
    return None

# Merge and write section headers based on which columns exist
for section_name, names in sections.items():
    start = end = None
    for header_name in names:
        col_found = find_col_for_header_name_in_output(header_name)
        if col_found:
            start = start or col_found
            end = col_found
    if start and end:
        ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        sec = ws.cell(row=2, column=start, value=section_name)
        sec.font = bold_font
        sec.alignment = center
        if "Planned" in section_name:
            sec.fill = planned_fill
        elif "Actual" in section_name:
            sec.fill = actual_fill
        elif "Deviation Days" in section_name:
            sec.fill = dev_fill
        elif "Total Deviation" in section_name:
            sec.fill = percent_fill
        border_merged(ws, f"{get_column_letter(start)}2:{get_column_letter(end)}2")

# -------------------------------------------------------
# STEP F: Style the header row (row 3) — set fills for recognized headers
# -------------------------------------------------------
# -------------------------------------------------------
# -------------------------------------------------------
# STEP F (FINAL): Color headers STRICTLY based on row-2
# merged section titles. Row-3 text is ignored completely.
# -------------------------------------------------------

# Map keywords → fill color
def get_section_fill(section_name):
    if not section_name:
        return None
    s = str(section_name).lower()

    if "planned" in s:
        return planned_fill
    if "actual" in s:
        return actual_fill
    if "deviation days" in s:
        return dev_fill
    if "total deviation" in s or "deviation %" in s or "dev%" in s:
        return percent_fill
    if "extra" in s:
        return extra_fill

    return None


# Clear header styling first
for c in range(1, ws.max_column + 1):
    cell = ws.cell(row=HEADER_OUTPUT_ROW, column=c)
    cell.font = bold_font
    cell.border = thin_border
    cell.alignment = center
    cell.fill = PatternFill()   # reset


#  Identify all merged section blocks in row 2 and color everything under them
colored = set()

for merged in ws.merged_cells.ranges:
    if merged.min_row == 2 and merged.max_row == 2:
        start_col = merged.min_col
        end_col = merged.max_col

        section_title = ws.cell(row=2, column=start_col).value
        fill = get_section_fill(section_title)

        if fill:
            for c in range(start_col, end_col + 1):
                ws.cell(row=HEADER_OUTPUT_ROW, column=c).fill = fill
                colored.add(c)


#  Handle the columns in row 2 that are NOT merged (single-cell sections)
for c in range(1, ws.max_column + 1):
    if c in colored:
        continue

    sec = ws.cell(row=2, column=c).value
    fill = get_section_fill(sec)

    if fill:
        ws.cell(row=HEADER_OUTPUT_ROW, column=c).fill = fill
        colored.add(c)

# DONE — NO MORE FALLBACKS BASED ON TEXT


# -------------------------------------------------------
# STEP G: CLEAN HEADINGS — remove prefixes (Planned_, Actual_, Dev_, Dev%_)
# -------------------------------------------------------

def clean_heading(name: str):
    """
    Converts:
       Planned_Apr25-Oct25 → Apr25-Oct25 P
       Actual_Apr25-Oct25  → Apr25-Oct25 A
       Dev_Apr25-Oct25     → Apr25-Oct25 D
       Dev%_Apr25-Oct25    → Apr25-Oct25
       Planned_Total       → Total P
       Actual_Total        → Total A
       Dev_Total           → Total D
       Dev%_Total          → Total
    """
    if name.startswith("Planned_"):
        suffix = name.replace("Planned_", "")
        return f"{suffix} P" if suffix != "Total" else "Total P"

    if name.startswith("Actual_"):
        suffix = name.replace("Actual_", "")
        return f"{suffix} A" if suffix != "Total" else "Total A"

    if name.startswith("Dev_"):
        suffix = name.replace("Dev_", "")
        return f"{suffix} D" if suffix != "Total" else "Total D"

    if name.startswith("Dev%_"):
        suffix = name.replace("Dev%_", "")
        return suffix if suffix != "Total" else "Total"

    return name


# Apply cleaned headings to row 3
for c in range(1, ws.max_column + 1):
    val = ws.cell(row=HEADER_OUTPUT_ROW, column=c).value
    if val:
        ws.cell(row=HEADER_OUTPUT_ROW, column=c).value = clean_heading(str(val))
        ws.cell(row=HEADER_OUTPUT_ROW, column=c).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


# -------------------------------------------------------
# STEP H: Data formatting (zero → blank) and borders for existing cells
# Use current ws.max_column so newly inserted columns are handled later.
# -------------------------------------------------------
for r in range(HEADER_OUTPUT_ROW + 1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        if cell.value == 0:
            cell.value = ""
        cell.border = thin_border

ws.freeze_panes = "A4"

# -------------------------------------------------------
# STEP I: EXTRA PERSONS COLUMNS — insert 3 columns at the end
# -------------------------------------------------------
extra_start = ws.max_column + 1
ws.insert_cols(extra_start, 3)

# Friendly labels for extra columns (placed on output header row)
ws.cell(row=HEADER_OUTPUT_ROW, column=extra_start, value=f"{group1_label}")
ws.cell(row=HEADER_OUTPUT_ROW, column=extra_start + 1, value=f"{group2_label}")
ws.cell(row=HEADER_OUTPUT_ROW, column=extra_start + 2, value="Total A\n(" + (group1_label + " + " + group2_label if group1_label else group2_label) + ")")

for c in range(extra_start, extra_start + 3):
    cc = ws.cell(row=HEADER_OUTPUT_ROW, column=c)
    cc.font = bold_font
    cc.alignment = center
    cc.border = thin_border
    cc.fill = extra_fill

# Merge red header on row 2 for these columns
ws.merge_cells(start_row=2, start_column=extra_start, end_row=2, end_column=extra_start + 2)
h = ws.cell(row=2, column=extra_start, value="Total Extra Persons")
h.font = bold_font
h.alignment = center
h.fill = extra_fill
border_merged(ws, f"{get_column_letter(extra_start)}2:{get_column_letter(extra_start+2)}2")

# -------------------------------------------------------
# STEP J: Compute Extra Persons using correct dynamic columns
# Group1 divisor = 24 ; Group2 divisor = 19
# We'll find actual column indexes of the dynamic headers in header_to_col_output
# -------------------------------------------------------
# Build a mapping of dynamic header names to column indices (if present)
col_P_G1 = header_to_col_output.get(group_keys['Planned_G1'])
col_P_G2 = header_to_col_output.get(group_keys['Planned_G2'])
col_A_G1 = header_to_col_output.get(group_keys['Actual_G1'])
col_A_G2 = header_to_col_output.get(group_keys['Actual_G2'])

# If exact dynamic names are not found, attempt fuzzy matches using label tokens
def fuzzy_find(col_map, token):
    if not token:
        return None
    for name, col in col_map.items():
        if token in name:
            return col
    return None

if not col_P_G1:
    col_P_G1 = fuzzy_find(header_to_col_output, group1_label)
if not col_P_G2:
    col_P_G2 = fuzzy_find(header_to_col_output, group2_label)
if not col_A_G1:
    col_A_G1 = fuzzy_find(header_to_col_output, group1_label)
if not col_A_G2:
    col_A_G2 = fuzzy_find(header_to_col_output, group2_label)

# Now compute row-wise extras
for r in range(HEADER_OUTPUT_ROW + 1, ws.max_row + 1):
    planned_g1 = ws.cell(row=r, column=col_P_G1).value if col_P_G1 else 0
    actual_g1 = ws.cell(row=r, column=col_A_G1).value if col_A_G1 else 0
    planned_g2 = ws.cell(row=r, column=col_P_G2).value if col_P_G2 else 0
    actual_g2 = ws.cell(row=r, column=col_A_G2).value if col_A_G2 else 0

    # ensure numeric values
    try:
        planned_g1 = float(planned_g1) if planned_g1 not in (None, "") else 0
    except:
        planned_g1 = 0
    try:
        actual_g1 = float(actual_g1) if actual_g1 not in (None, "") else 0
    except:
        actual_g1 = 0
    try:
        planned_g2 = float(planned_g2) if planned_g2 not in (None, "") else 0
    except:
        planned_g2 = 0
    try:
        actual_g2 = float(actual_g2) if actual_g2 not in (None, "") else 0
    except:
        actual_g2 = 0

    # Group1 extra (divide by 24) if actual > planned
    if actual_g1 > planned_g1:
        extra_g1 = round((actual_g1 - planned_g1) / 24, 2)
    else:
        extra_g1 = ""

    # Group2 extra (divide by 19) if actual > planned
    if actual_g2 > planned_g2:
        extra_g2 = round((actual_g2 - planned_g2) / 19, 2)
    else:
        extra_g2 = ""

    if extra_g1 != "" or extra_g2 != "":
        total_extra = round((extra_g1 or 0) + (extra_g2 or 0), 2)
    else:
        total_extra = ""

    ws.cell(row=r, column=extra_start, value=extra_g1)
    ws.cell(row=r, column=extra_start + 1, value=extra_g2)
    ws.cell(row=r, column=extra_start + 2, value=total_extra)

# Add borders for final extra columns
for r in range(HEADER_OUTPUT_ROW, ws.max_row + 1):
    for c in range(extra_start, extra_start + 3):
        ws.cell(row=r, column=c).border = thin_border

# -------------------------------------------------------
# SAVE
# -------------------------------------------------------
wb.save(file_path)
print("Formatting + sorting + dynamic month grouping + headings + extra persons updated (Option A).")  