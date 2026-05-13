from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import sys

# === Load file ===
file_path = sys.argv[1]
wb = load_workbook(file_path)
ws = wb.active

# === Styles ===
bold_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

# Header colors
planned_fill = PatternFill(start_color="FFF2CC", fill_type="solid")   # light yellow
actual_fill = PatternFill(start_color="CFE2F3", fill_type="solid")    # light blue
deviation_fill = PatternFill(start_color="F9CB9C", fill_type="solid") # orange
percent_fill = PatternFill(start_color="D9EAD3", fill_type="solid")   # green

# Deviation colors
positive_fill = PatternFill(start_color="C6EFCE", fill_type="solid")  # light green
negative_fill = PatternFill(start_color="F4CCCC", fill_type="solid")  # light red

# === Add Sr. No if missing ===
if ws.cell(row=1, column=1).value != "SrNo" and ws.cell(row=1, column=1).value != "Sr. No.":
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="Sr. No.").font = bold_font
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
        row[0].value = i

# === Build header map ===
header_map = {}
for cell in ws[1]:
    header_value = str(cell.value or "").strip()
    header_map[cell.column] = header_value

    cell.font = bold_font
    cell.alignment = center
    cell.border = thin_border

    # Header color
    if "Planned Days" in header_value:
        cell.fill = planned_fill
    elif "Actual Days" in header_value:
        cell.fill = actual_fill
    elif "Deviation Days" in header_value:
        cell.fill = deviation_fill
    elif "Total Deviation %" in header_value:
        cell.fill = percent_fill

# === Apply formatting to rows ===
for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in r:
        cell.border = thin_border
        header = header_map.get(cell.column, "")

        # Bold parent rows (Parent column has text)
        if header == "Parent" and cell.value not in ["", None]:
            cell.font = bold_font

        # Highlight positive / negative deviation %
        if header == "Total Deviation %":
            try:
                val = float(cell.value)
                if val > 0:
                    cell.fill = positive_fill
                elif val < 0:
                    cell.fill = negative_fill
            except:
                pass

# === Column Widths ===
for col in ws.columns:
    col_letter = col[0].column_letter
    header = str(ws.cell(row=1, column=col[0].column).value or "")

    if header == "Sr. No.":
        ws.column_dimensions[col_letter].width = 7
    elif header == "Parent" or header == "PM":
        ws.column_dimensions[col_letter].width = 20
    else:
        ws.column_dimensions[col_letter].width = 15

# Freeze header
ws.freeze_panes = "A2"

# === Save file ===
wb.save(file_path)
